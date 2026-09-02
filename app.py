#region IMPORTS

from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

#endregion IMPORTS


#region APP CONFIGURATION

APP_TITLE = "Customer Scoring - Web V15"
BASE_DIR = Path(__file__).parent
DATA_FILE = BASE_DIR / "Customer_Scoring.xlsx"
ASSETS_DIR = BASE_DIR / "assets"

SPINMASTER_LOGO_PATH = ASSETS_DIR / "SpinMaster-Logo-CMYK.png"
PRIMAL_HATCH_IMAGE_PATH = ASSETS_DIR / "PrimalHatch_Jurassic_DinoFreedom-2025.png"
PAW_HEADER_IMAGE_PATH = ASSETS_DIR / "PAW_CSG25_Grp_005_CGI.jpg"

MAX_TOTAL = 15

MONTH_WEEK_MAPPING = {
    1:  [1, 2, 3, 4],
    2:  [5, 6, 7, 8],
    3:  [9, 10, 11, 12, 13],
    4:  [14, 15, 16, 17],
    5:  [18, 19, 20, 21],
    6:  [22, 23, 24, 25, 26],
    7:  [27, 28, 29, 30],
    8:  [31, 32, 33, 34],
    9:  [35, 36, 37, 38, 39],
    10: [40, 41, 42, 43],
    11: [44, 45, 46, 47],
    12: [48, 49, 50, 51, 52],
}

MONTH_LABELS = [
    "1 - January",
    "2 - February",
    "3 - March",
    "4 - April",
    "5 - May",
    "6 - June",
    "7 - July",
    "8 - August",
    "9 - September",
    "10 - October",
    "11 - November",
    "12 - December",
]

CATEGORIES = [
    (
        "Timeliness",
        "How reliably the customer sends the data on time.\n\n"
        "0 = Missing / seriously late\n"
        "1 = Often late\n"
        "2 = Mostly on time, minor delays\n"
        "3 = On time",
    ),
    (
        "Layout Consistency",
        "How stable and reusable the incoming file layout is.\n\n"
        "0 = Major layout changes / difficult to process\n"
        "1 = Frequent layout changes\n"
        "2 = Mostly consistent, small changes\n"
        "3 = Fully consistent",
    ),
    (
        "Data Completeness",
        "Whether the required POS / inventory / product information is present.\n\n"
        "0 = Major information missing\n"
        "1 = Several missing elements\n"
        "2 = Mostly complete, minor gaps\n"
        "3 = Complete",
    ),
    (
        "Material Mapping",
        "How cleanly customer materials can be matched to the correct #600.\n\n"
        "0 = Major mapping problems\n"
        "1 = Frequent manual mapping required\n"
        "2 = Minor mapping issues\n"
        "3 = Clean / accurate mapping",
    ),
    (
        "Manual Effort",
        "How much analyst work is needed before the data can be used.\n\n"
        "0 = Very high manual effort\n"
        "1 = High manual effort\n"
        "2 = Some manual work needed\n"
        "3 = Minimal manual effort",
    ),
]

DEFAULT_CUSTOMERS = {
    "UK": ["Amazon UK", "Argos", "ASDA", "B&M", "ENTERTAINER", "Sainsburys", "Smyths UK"],
    "FR": ["Amazon FR", "Auchan FR", "Carrefour FR", "Cultura", "Distritoys", "Fnac FR", "Joueclub & Ludendo", "Leclerc", "Maxitoys", "Smyths FR"],
    "GAS": ["Amazon DE", "Wave", "Karstad", "Kaufland", "Mueller", "Otto", "Rofu", "Rossmann", "Smyths DE", "Thalia", "Interspar", "Migros"],
    "IBER": ["Amazon ES", "Alcampo", "Carrefour ES", "El Corte Ingles", "Fnac ES", "Toy Planet", "Toys r us", "El Corte Ingles Portugal", "Pingo Doce", "Sonae"],
    "IT": ["Amazon IT", "Prenatal IT"],
    "BNL": ["Amazon NL", "Amazon BE", "Bol.com", "LOBBES", "Intertoys", "Colruyt", "Dreamland", "Wehkamp"],
    "GR": ["Enarxis", "Jumbo", "Max Stores", "Moustakas", "Perfect Toys", "Retail World"],
    "CEE": ["Auchan HU", "Modell & Hobby", "Regio", "Spar", "Carrefour PL", "Amazon PL", "Smyk", "Auchan RO", "Carrefour RO", "Noriel/Intertoy", "Peaktoys", "Allegro", "Alza", "Sparkys", "Alltoys", "Framee", "TESCO CEE"],
    "NORDICS": ["Lekia", "SG"],
}

DEFAULT_FREQUENCY = "Weekly"

MONTHLY_CUSTOMERS = {
    "Cultura",
    "Carrefour FR",
    "Leclerc",
    "Interspar",
    "Kaufland",
    "Modell & Hobby",
    "Carrefour PL",
    "Auchan RO",
    "Carrefour RO",
    "Noriel/Intertoy",
    "Peaktoys",
}

CATEGORY_COLORS = {
    "Timeliness": "#20a8f0",
    "Layout Consistency": "#f04a34",
    "Data Completeness": "#9a5de8",
    "Material Mapping": "#79c934",
    "Manual Effort": "#f39a22",
}

#endregion APP CONFIGURATION


#region REPORTING HELPERS

def parse_month_value(value):
    text = str(value or "").strip()
    if not text:
        raise ValueError("Month is empty.")
    month = int(text.split("-", 1)[0].strip())
    if month not in MONTH_WEEK_MAPPING:
        raise ValueError("Month must be between 1 and 12.")
    return month


def weeks_for_month(month):
    return list(MONTH_WEEK_MAPPING[month])


def month_for_week(week):
    for month, weeks in MONTH_WEEK_MAPPING.items():
        if week in weeks:
            return month
    return 12


def get_customer_frequency(region, customer):
    if region == "GR":
        return "Monthly"
    if customer in MONTHLY_CUSTOMERS:
        return "Monthly"
    return "Weekly"


def parse_custom_weeks(year, text):
    text = (text or "").strip()
    if not text:
        return []

    weeks = []

    for part in text.split(","):
        part = part.strip()
        if not part:
            continue

        if "-" in part:
            a, b = part.split("-", 1)
            start_week = int(a.strip())
            end_week = int(b.strip())

            if start_week > end_week:
                start_week, end_week = end_week, start_week

            for w in range(start_week, end_week + 1):
                datetime.fromisocalendar(year, w, 1)
                weeks.append(w)
        else:
            w = int(part)
            datetime.fromisocalendar(year, w, 1)
            weeks.append(w)

    return [(year, w) for w in sorted(set(weeks))]


#endregion REPORTING HELPERS


#region EXCEL DATA STORAGE

def format_workbook(wb):
    fill = PatternFill("solid", fgColor="243447")
    hdr_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        for c in ws[1]:
            c.fill = fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row, col).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 11), 30)


def ensure_workbook():
    if DATA_FILE.exists():
        sync_customer_master()
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    ws.append([
        "Year", "Week", "Region", "Customer", "Frequency",
        "Timeliness", "Layout Consistency", "Data Completeness",
        "Material Mapping", "Manual Effort", "Total", "Life %", "Saved At"
    ])

    cws = wb.create_sheet("Customers")
    cws.append(["Region", "Customer", "Frequency"])

    for region, customers in DEFAULT_CUSTOMERS.items():
        for customer in customers:
            cws.append([region, customer, get_customer_frequency(region, customer)])

    format_workbook(wb)
    wb.save(DATA_FILE)
    wb.close()


def sync_customer_master():
    if not DATA_FILE.exists():
        return

    wb = load_workbook(DATA_FILE)

    if "Customers" not in wb.sheetnames:
        ws = wb.create_sheet("Customers")
        ws.append(["Region", "Customer", "Frequency"])
    else:
        ws = wb["Customers"]

    if ws.cell(1, 3).value != "Frequency":
        ws.cell(1, 3).value = "Frequency"

    existing = {}

    for row_idx in range(2, ws.max_row + 1):
        region = ws.cell(row_idx, 1).value
        customer = ws.cell(row_idx, 2).value

        if region and customer:
            region = str(region)
            customer = str(customer)
            existing[(region, customer)] = row_idx
            ws.cell(row_idx, 3).value = get_customer_frequency(region, customer)

    for region, customers in DEFAULT_CUSTOMERS.items():
        for customer in customers:
            if (region, customer) not in existing:
                ws.append([region, customer, get_customer_frequency(region, customer)])

    format_workbook(wb)
    wb.save(DATA_FILE)
    wb.close()


def replace_data_file(uploaded_file):
    DATA_FILE.write_bytes(uploaded_file.getvalue())
    ensure_workbook()


def load_customer_master():
    ensure_workbook()

    wb = load_workbook(DATA_FILE, data_only=True)
    ws = wb["Customers"]
    data = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        region, customer = row[0], row[1]

        if not region or not customer:
            continue

        region = str(region)
        customer = str(customer)

        data[region].append({
            "customer": customer,
            "frequency": get_customer_frequency(region, customer),
        })

    wb.close()
    return dict(data)


def get_saved_customers_for_period(year, weeks):
    ensure_workbook()

    if isinstance(weeks, int):
        weeks = [weeks]

    weeks = set(weeks)

    wb = load_workbook(DATA_FILE, data_only=True)
    ws = wb["Scores"]
    saved_by_customer = defaultdict(set)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        row_year = row[0]
        row_week = row[1]
        customer = row[3]

        if row_year == year and row_week in weeks and customer:
            saved_by_customer[str(customer)].add(int(row_week))

    wb.close()

    # For a monthly selection, show complete only if the customer has every week.
    return {
        customer
        for customer, saved_weeks in saved_by_customer.items()
        if weeks.issubset(saved_weeks)
    }


def get_last_saved_period():
    ensure_workbook()

    wb = load_workbook(DATA_FILE, data_only=True)
    ws = wb["Scores"]
    periods = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], int) and isinstance(row[1], int):
            periods.append((row[0], row[1]))

    wb.close()

    if periods:
        return max(periods, key=lambda p: datetime.fromisocalendar(p[0], p[1], 1))

    reporting_date = datetime.now() - timedelta(weeks=1)
    iso = reporting_date.isocalendar()
    return iso.year, iso.week


def migrate_old_scores_if_needed(wb):
    ws = wb["Scores"]
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]

    if "Frequency" in headers:
        return

    old_rows = list(ws.iter_rows(min_row=2, values_only=True))
    old_index = wb.sheetnames.index("Scores")
    wb.remove(ws)

    nws = wb.create_sheet("Scores", old_index)
    nws.append([
        "Year", "Week", "Region", "Customer", "Frequency",
        "Timeliness", "Layout Consistency", "Data Completeness",
        "Material Mapping", "Manual Effort", "Total", "Life %", "Saved At"
    ])

    for row in old_rows:
        if not row or row[0] is None:
            continue

        region = str(row[2])
        customer = str(row[3])
        freq = get_customer_frequency(region, customer)

        nws.append([
            row[0], row[1], row[2], row[3], freq,
            row[4], row[5], row[6], row[7], row[8],
            row[9], row[10], row[11]
        ])


def save_batch_to_excel(batch, year, week):
    ensure_workbook()

    wb = load_workbook(DATA_FILE)
    migrate_old_scores_if_needed(wb)
    ws = wb["Scores"]

    existing = {}

    for r in range(2, ws.max_row + 1):
        y = ws.cell(r, 1).value
        w = ws.cell(r, 2).value
        customer = ws.cell(r, 4).value

        if y is not None and w is not None and customer:
            existing[(int(y), int(w), str(customer))] = r

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for customer, item in batch.items():
        s = item["scores"]
        total = sum(s.values())

        values = [
            year,
            week,
            item["region"],
            customer,
            item["frequency"],
            s["Timeliness"],
            s["Layout Consistency"],
            s["Data Completeness"],
            s["Material Mapping"],
            s["Manual Effort"],
            total,
            total / MAX_TOTAL,
            stamp,
        ]

        key = (year, week, customer)

        if key in existing:
            rr = existing[key]
            for cc, value in enumerate(values, start=1):
                ws.cell(rr, cc).value = value
        else:
            ws.append(values)

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 12).number_format = "0%"

    format_workbook(wb)
    wb.save(DATA_FILE)
    wb.close()


def load_all_scores():
    ensure_workbook()

    wb = load_workbook(DATA_FILE, data_only=True)
    ws = wb["Scores"]
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    has_freq = "Frequency" in headers
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        if has_freq:
            freq = row[4] or DEFAULT_FREQUENCY
            offset = 1
        else:
            freq = get_customer_frequency(str(row[2]), str(row[3]))
            offset = 0

        records.append({
            "year": row[0],
            "week": row[1],
            "region": row[2],
            "customer": row[3],
            "frequency": str(freq).title(),
            "scores": {
                "Timeliness": row[4 + offset],
                "Layout Consistency": row[5 + offset],
                "Data Completeness": row[6 + offset],
                "Material Mapping": row[7 + offset],
                "Manual Effort": row[8 + offset],
            },
            "total": row[9 + offset],
            "life": row[10 + offset],
        })

    wb.close()
    return records


def workbook_bytes():
    ensure_workbook()
    return DATA_FILE.read_bytes()


#endregion EXCEL DATA STORAGE


#region VISUAL DATA PREPARATION

def periods_back(year, week, count=5):
    current = datetime.fromisocalendar(year, week, 1)
    result = []

    for offset in range(count - 1, -1, -1):
        d = current - timedelta(weeks=offset)
        iso = d.isocalendar()
        result.append((iso.year, iso.week))

    return result


def prepare_visual_rows(
    year,
    week,
    mode,
    region_filter,
    frequency_filter,
    custom_periods=None,
):
    records = load_all_scores()

    if mode == "Latest Week":
        periods = [(year, week)]
    elif mode == "Last 5 Weeks":
        periods = periods_back(year, week, 5)
    else:
        periods = custom_periods or []

    wanted = set(periods)

    filtered = []

    for r in records:
        if (r["year"], r["week"]) not in wanted:
            continue

        if region_filter != "All Regions" and r["region"] != region_filter:
            continue

        if frequency_filter != "All" and r["frequency"] != frequency_filter:
            continue

        filtered.append(r)

    if mode == "Latest Week":
        rows = []

        for r in filtered:
            life = float(r["life"] or 0)

            rows.append({
                "region": r["region"],
                "customer": r["customer"],
                "frequency": r["frequency"],
                "life": life,
                "scores": {
                    key: float(value or 0)
                    for key, value in r["scores"].items()
                },
                "available": 1,
                "selected_count": 1,
            })

        rows.sort(key=lambda x: (-x["life"], x["customer"]))
        return rows, periods

    grouped = {}

    for r in filtered:
        key = (r["region"], r["customer"], r["frequency"])

        grouped.setdefault(
            key,
            {
                "region": r["region"],
                "customer": r["customer"],
                "frequency": r["frequency"],
                "records": {},
            },
        )

        grouped[key]["records"][(r["year"], r["week"])] = r

    rows = []

    for g in grouped.values():
        available_records = [
            g["records"][p]
            for p in periods
            if p in g["records"]
        ]

        if not available_records:
            continue

        avg_life = (
            sum(float(r["life"] or 0) for r in available_records)
            / len(available_records)
        )

        avg_scores = {}

        for category, _tooltip in CATEGORIES:
            values = [
                float(r["scores"].get(category) or 0)
                for r in available_records
                if r["scores"].get(category) is not None
            ]

            avg_scores[category] = (
                sum(values) / len(values)
                if values else 0
            )

        rows.append({
            "region": g["region"],
            "customer": g["customer"],
            "frequency": g["frequency"],
            "life": avg_life,
            "scores": avg_scores,
            "available": len(available_records),
            "selected_count": len(periods),
        })

    rows.sort(key=lambda x: (-x["life"], x["customer"]))
    return rows, periods


#endregion VISUAL DATA PREPARATION


#region VISUAL RENDERING

def life_color(percent):
    if percent >= 90:
        return "#49d70f"
    if percent >= 75:
        return "#a8dc12"
    if percent >= 50:
        return "#f6c313"
    if percent >= 25:
        return "#ef8014"
    return "#e63824"


def font(size, bold=False):
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold
        else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/arialbd.ttf" if bold
        else "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/segoeuib.ttf" if bold
        else "C:/Windows/Fonts/segoeui.ttf",
    ]

    for p in candidates:
        try:
            return ImageFont.truetype(p, size)
        except Exception:
            pass

    return ImageFont.load_default()


def centered(draw, box, text, fnt, fill):
    x1, y1, x2, y2 = box
    b = draw.textbbox((0, 0), text, font=fnt)
    tw = b[2] - b[0]
    th = b[3] - b[1]

    draw.text(
        (x1 + (x2 - x1 - tw) / 2, y1 + (y2 - y1 - th) / 2 - 1),
        text,
        font=fnt,
        fill=fill,
    )


def draw_score_dots(draw, x, y, score, color):
    score = max(0.0, min(3.0, float(score or 0)))

    for i in range(3):
        cx = x + i * 22
        left, top, right, bottom = cx - 7, y - 7, cx + 7, y + 7
        local = max(0.0, min(1.0, score - i))

        draw.ellipse(
            (left, top, right, bottom),
            fill="#18232d",
            outline="#6d7a84",
            width=2,
        )

        if local >= 0.999:
            draw.ellipse(
                (left, top, right, bottom),
                fill=color,
                outline="#d9f4ff",
                width=2,
            )
        elif local > 0:
            fill_right = left + int((right - left) * local)

            draw.rectangle(
                (left, top, fill_right, bottom),
                fill=color,
            )

            draw.ellipse(
                (left, top, right, bottom),
                outline="#d9f4ff",
                width=2,
            )


def render_ranking(rows, year, week, mode, periods):
    width = 1160
    top_h, header_h, row_h, footer_h = 155, 62, 48, 18
    count = max(1, len(rows))
    height = top_h + header_h + row_h * count + footer_h

    img = Image.new("RGB", (width, height), "#07111d")
    d = ImageDraw.Draw(img)

    title = font(36, True)
    sub = font(14, True)
    hdr = font(12, True)
    body = font(13, True)
    small = font(9, False)
    pctf = font(15, True)
    rankf = font(18, True)

    d.rectangle((0, 0, width, top_h), fill="#06131f")

    # Optional logo if user later adds it to /assets.
    if SPINMASTER_LOGO_PATH.exists():
        try:
            logo = Image.open(SPINMASTER_LOGO_PATH).convert("RGBA")
            logo.thumbnail((125, 60), Image.LANCZOS)
            img.paste(logo, (22, 25), logo)
        except Exception:
            pass

    if PAW_HEADER_IMAGE_PATH.exists():
        try:
            paw = Image.open(PAW_HEADER_IMAGE_PATH).convert("RGBA")
            paw.thumbnail((145, 80), Image.LANCZOS)
            img.paste(paw, (width - paw.width - 20, 18), paw)
        except Exception:
            pass

    centered(d, (180, 16, width - 180, 66), "CUSTOMER RANKING", title, "#ffc928")

    if mode == "Latest Week":
        view_label = "LATEST WEEK"
    elif mode == "Last 5 Weeks":
        view_label = "LAST 5 WEEKS"
    else:
        view_label = "CUSTOM PERIOD"

    centered(
        d,
        (220, 70, width - 220, 101),
        f"{view_label}  •  W{week} / {year}",
        sub,
        "#f3f7fb",
    )

    if mode == "Latest Week":
        summary = ""
    else:
        summary = "AVERAGE OF AVAILABLE SCORES • " + "  ".join(
            f"W{w}" for _, w in periods
        )

    if summary:
        centered(
            d,
            (170, 108, width - 170, 136),
            summary,
            small,
            "#c7d8e5",
        )

    x_rank, w_rank = 12, 60
    x_customer, w_customer = 72, 205
    x_health, w_health = 277, 285
    x_cat, cat_w = 562, 116
    hy1, hy2 = top_h, top_h + header_h

    d.rectangle((15, hy1, width - 15, hy2), outline="#2f6f8f", width=2)
    centered(d, (x_rank, hy1, x_rank + w_rank, hy2), "RANK", hdr, "#f3f7fb")
    centered(d, (x_customer, hy1, x_customer + w_customer, hy2), "CUSTOMER", hdr, "#f3f7fb")
    centered(d, (x_health, hy1, x_health + w_health, hy2), "OVERALL HEALTH", hdr, "#f3f7fb")

    cat_short = [
        ("Timeliness", "TIMELINESS"),
        ("Layout Consistency", "LAYOUT"),
        ("Data Completeness", "DATA"),
        ("Material Mapping", "MAPPING"),
        ("Manual Effort", "EFFORT"),
    ]

    for i, (key, label) in enumerate(cat_short):
        x1 = x_cat + i * cat_w
        x2 = x_cat + (i + 1) * cat_w
        d.rectangle((x1, hy1, x2, hy2), outline="#2f6f8f", width=2)
        centered(d, (x1 + 3, hy1, x2 - 3, hy2), label, small, CATEGORY_COLORS[key])

    start_y = hy2

    if not rows:
        centered(
            d,
            (0, start_y, width, start_y + row_h),
            "NO SCORES MATCH THE CURRENT FILTERS",
            body,
            "#9eb2c2",
        )

    for idx, r in enumerate(rows, start=1):
        y1 = start_y + (idx - 1) * row_h
        y2 = start_y + idx * row_h

        d.rectangle(
            (15, y1, width - 15, y2),
            fill="#091825" if idx % 2 else "#0c1e2d",
        )

        for x in [
            x_rank, x_customer, x_health, x_cat,
            x_cat + cat_w, x_cat + 2 * cat_w,
            x_cat + 3 * cat_w, x_cat + 4 * cat_w,
            x_cat + 5 * cat_w,
        ]:
            d.line((x, y1, x, y2), fill="#2f6f8f", width=1)

        d.line((15, y2, width - 15, y2), fill="#2f6f8f", width=1)

        centered(
            d,
            (x_rank, y1, x_rank + w_rank, y2),
            str(idx),
            rankf,
            "#ffc928" if idx == 1 else "#dbe8f0",
        )

        d.text(
            (x_customer + 10, y1 + 6),
            r["customer"],
            font=body,
            fill="#f3f7fb",
        )

        d.text(
            (x_customer + 10, y1 + 27),
            f'{r["region"]} • {r["frequency"]}',
            font=small,
            fill="#9eb2c2",
        )

        life_pct = max(0, min(100, r["life"] * 100))

        bx1 = x_health + 10
        by1 = y1 + 10
        bx2 = x_health + 205
        by2 = y1 + 32

        d.rounded_rectangle(
            (bx1, by1, bx2, by2),
            radius=6,
            fill="#061018",
            outline="#8cc8df",
            width=2,
        )

        fw = int((bx2 - bx1 - 6) * life_pct / 100)

        if fw > 0:
            d.rounded_rectangle(
                (bx1 + 3, by1 + 3, bx1 + 3 + fw, by2 - 3),
                radius=4,
                fill=life_color(life_pct),
            )

        d.text(
            (x_health + 218, y1 + 11),
            f"{life_pct:.0f}%",
            font=pctf,
            fill=life_color(life_pct),
        )

        scores = r["scores"] or {}

        if mode != "Latest Week":
            d.text(
                (x_health + 10, y1 + 34),
                f'Based on {r["available"]}/{r["selected_count"]} selected week(s)',
                font=font(10),
                fill="#9eb2c2",
            )

        for i, (key, _) in enumerate(cat_short):
            x1 = x_cat + i * cat_w
            score = float(scores.get(key, 0) or 0)

            draw_score_dots(
                d,
                x1 + 30,
                y1 + 20,
                score,
                CATEGORY_COLORS[key],
            )

            if mode != "Latest Week":
                centered(
                    d,
                    (x1 + 3, y1 + 30, x1 + cat_w - 3, y2 - 1),
                    f"{score:.1f}/3",
                    font(10, True),
                    CATEGORY_COLORS[key],
                )

    return img


def image_to_png_bytes(image):
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


#endregion VISUAL RENDERING


#region STREAMLIT UI HELPERS

def init_session_state():
    reporting_date = datetime.now() - timedelta(weeks=1)
    iso = reporting_date.isocalendar()

    defaults = {
        "pending_scores": {},
        "year": int(iso.year),
        "week": int(iso.week),
        "month": month_for_week(int(iso.week)),
        "region": "UK",
        "customer": "",
        "customer_search": "",
        "score_Timeliness": 3,
        "score_Layout Consistency": 3,
        "score_Data Completeness": 3,
        "score_Material Mapping": 3,
        "score_Manual Effort": 3,
    }

    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def current_score_dict():
    return {
        category: int(st.session_state[f"score_{category}"])
        for category, _ in CATEGORIES
    }


def pending_table_rows():
    rows = []

    for customer, item in st.session_state.pending_scores.items():
        s = item["scores"]

        rows.append({
            "Region": item["region"],
            "Customer": customer,
            "Frequency": item["frequency"],
            "Period": item["period_label"],
            "Timeliness": s["Timeliness"],
            "Layout": s["Layout Consistency"],
            "Data": s["Data Completeness"],
            "Mapping": s["Material Mapping"],
            "Effort": s["Manual Effort"],
            "Total": sum(s.values()),
        })

    return rows


def selected_period_for_frequency(frequency):
    year = int(st.session_state.year)

    if frequency == "Monthly":
        month = int(st.session_state.month)
        return year, weeks_for_month(month), f"M{month}"

    week = int(st.session_state.week)
    datetime.fromisocalendar(year, week, 1)
    return year, [week], f"W{week}"


def selected_completed_customers(frequency):
    year = int(st.session_state.year)

    if frequency == "Monthly":
        weeks = weeks_for_month(int(st.session_state.month))
    else:
        weeks = [int(st.session_state.week)]

    return get_saved_customers_for_period(year, weeks)


def add_current_customer_to_batch(region, customer):
    if not customer:
        st.warning("Select a customer first.")
        return

    frequency = get_customer_frequency(region, customer)
    year, target_weeks, period_label = selected_period_for_frequency(frequency)

    st.session_state.pending_scores[customer] = {
        "region": region,
        "frequency": frequency,
        "scores": current_score_dict(),
        "year": year,
        "target_weeks": target_weeks,
        "period_label": period_label,
    }

    st.success(f"{customer} added to pending batch.")


def save_pending_batch():
    if not st.session_state.pending_scores:
        st.info("The pending batch is empty.")
        return False

    grouped = {}

    for customer, item in st.session_state.pending_scores.items():
        year = item["year"]

        for week in item["target_weeks"]:
            key = (year, week)
            grouped.setdefault(key, {})

            grouped[key][customer] = {
                "region": item["region"],
                "frequency": item["frequency"],
                "scores": item["scores"],
            }

    for (year, week), week_batch in grouped.items():
        save_batch_to_excel(week_batch, year, week)

    customer_count = len(st.session_state.pending_scores)
    period_count = len(grouped)
    st.session_state.pending_scores = {}

    st.success(
        f"{customer_count} customer score(s) saved. "
        f"Excel week-periods written: {period_count}."
    )

    return True


def inject_css():
    st.markdown(
        """
        <style>
        .block-container {
            max-width: 1450px;
            padding-top: 1.2rem;
            padding-bottom: 2rem;
        }

        .sm-header {
            background: #243447;
            border-radius: 12px;
            padding: 16px 20px;
            margin-bottom: 14px;
            color: white;
        }

        .sm-header h1 {
            margin: 0;
            font-size: 30px;
            line-height: 1.1;
        }

        .sm-header span {
            color: #9ec6e8;
            font-size: 13px;
            font-weight: 700;
        }

        .section-title {
            font-size: 18px;
            font-weight: 800;
            color: #243447;
            margin-top: 4px;
            margin-bottom: 8px;
        }

        div[data-testid="stMetric"] {
            background: white;
            border: 1px solid #e1e7ec;
            border-radius: 12px;
            padding: 8px 12px;
        }

        div[data-testid="stDataFrame"] {
            border: 1px solid #dfe5ea;
            border-radius: 10px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


#endregion STREAMLIT UI HELPERS


#region STREAMLIT APP

st.set_page_config(
    page_title="Customer Ranking",
    page_icon="🎮",
    layout="wide",
)

inject_css()
ensure_workbook()
init_session_state()

header_logo, header_title, header_paw = st.columns([0.7, 4.8, 0.9], vertical_alignment="center")

with header_logo:
    if SPINMASTER_LOGO_PATH.exists():
        st.image(str(SPINMASTER_LOGO_PATH), width=92)

with header_title:
    st.markdown(
        """
        <div class="sm-header">
            <h1>CUSTOMER SCORING</h1>
            <span>WEB V15 • CUSTOMER RANKING</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

with header_paw:
    if PAW_HEADER_IMAGE_PATH.exists():
        st.image(str(PAW_HEADER_IMAGE_PATH), width=82)

with st.sidebar:
    st.header("Data file")

    st.caption(
        "The app reads and updates the bundled Customer_Scoring.xlsx. "
        "Download a backup whenever needed. For the final multi-user version, "
        "we will connect this same logic to the SharePoint master workbook."
    )

    uploaded_db = st.file_uploader(
        "Load an existing Customer_Scoring.xlsx",
        type=["xlsx"],
        key="database_upload",
    )

    if uploaded_db is not None:
        if st.button("Use uploaded workbook", use_container_width=True):
            try:
                replace_data_file(uploaded_db)
                st.success("Workbook loaded.")
                st.rerun()
            except Exception as exc:
                st.error(f"Could not load workbook: {exc}")

    st.download_button(
        "Download current workbook",
        data=workbook_bytes(),
        file_name="Customer_Scoring.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.divider()
    st.caption("Optional images can later be added to an `assets` folder in GitHub.")

tab_score, tab_rank = st.tabs(["📝 Score Customers", "🏆 Customer Ranking"])


# -------------------- SCORE CUSTOMERS --------------------
with tab_score:
    master = load_customer_master()
    regions = list(master.keys())

    if st.session_state.region not in regions and regions:
        st.session_state.region = regions[0]

    left, middle, right = st.columns([1.05, 1.05, 1.8], gap="large")

    with left:
        st.markdown('<div class="section-title">Selection</div>', unsafe_allow_html=True)

        period_a, period_b = st.columns(2)

        with period_a:
            st.number_input(
                "Year",
                min_value=2020,
                max_value=2100,
                step=1,
                key="year",
            )

        with period_b:
            st.number_input(
                "Week",
                min_value=1,
                max_value=53,
                step=1,
                key="week",
            )

        region = st.selectbox(
            "Region",
            options=regions,
            key="region",
        )

        st.text_input(
            "Search customer",
            key="customer_search",
            placeholder="Type part of customer name...",
        )

        entries = master.get(region, [])
        search_text = st.session_state.customer_search.strip().lower()

        visible_entries = [
            e for e in entries
            if search_text in e["customer"].lower()
        ]

        # Determine selection frequency before period field.
        customer_names = [e["customer"] for e in visible_entries]

        if st.session_state.customer not in customer_names:
            st.session_state.customer = customer_names[0] if customer_names else ""

        customer = st.selectbox(
            "Customer",
            options=customer_names if customer_names else [""],
            key="customer",
        )

        frequency = (
            get_customer_frequency(region, customer)
            if customer else "Weekly"
        )

        if frequency == "Monthly":
            st.selectbox(
                "Reporting month",
                options=list(range(1, 13)),
                format_func=lambda m: MONTH_LABELS[m - 1],
                key="month",
            )
            selected_weeks = weeks_for_month(int(st.session_state.month))
            st.caption(
                "Monthly customer • score will be written to "
                + ", ".join(f"W{w}" for w in selected_weeks)
            )
        else:
            st.caption(f"Weekly customer • score will be written to W{int(st.session_state.week)}")

        completed = selected_completed_customers(frequency)

        if customer:
            if customer in st.session_state.pending_scores:
                st.info("🟦 This customer is already in the pending batch.")
            elif customer in completed:
                st.success("✅ Already saved for the selected period.")
            else:
                st.caption("⬜ Not scored yet for the selected period.")

        st.caption("W = Weekly • M = Monthly")

    with middle:
        st.markdown('<div class="section-title">Score Customer</div>', unsafe_allow_html=True)
        st.caption("Choose a score from 0 to 3 for each category.")

        for category, tooltip in CATEGORIES:
            st.radio(
                category,
                options=[0, 1, 2, 3],
                horizontal=True,
                key=f"score_{category}",
                help=tooltip,
            )

        add_col, reset_col = st.columns(2)

        with add_col:
            if st.button(
                "ADD TO BATCH",
                type="primary",
                use_container_width=True,
            ):
                try:
                    add_current_customer_to_batch(region, customer)
                except Exception as exc:
                    st.error(str(exc))

        with reset_col:
            if st.button("RESET 3/3", use_container_width=True):
                for category, _ in CATEGORIES:
                    st.session_state[f"score_{category}"] = 3
                st.rerun()

        if PRIMAL_HATCH_IMAGE_PATH.exists():
            try:
                st.image(str(PRIMAL_HATCH_IMAGE_PATH), width=150)
            except Exception:
                pass

    with right:
        st.markdown('<div class="section-title">Pending Batch</div>', unsafe_allow_html=True)

        rows = pending_table_rows()

        if rows:
            st.dataframe(
                rows,
                use_container_width=True,
                hide_index=True,
            )

            remove_customer = st.selectbox(
                "Pending customer to edit/remove",
                options=[""] + list(st.session_state.pending_scores.keys()),
            )

            b1, b2, b3 = st.columns(3)

            with b1:
                if st.button("LOAD SCORES", use_container_width=True, disabled=not remove_customer):
                    item = st.session_state.pending_scores[remove_customer]
                    st.session_state.region = item["region"]
                    st.session_state.customer = remove_customer
                    st.session_state.year = item["year"]

                    if item["frequency"] == "Monthly":
                        st.session_state.month = month_for_week(item["target_weeks"][0])
                    else:
                        st.session_state.week = item["target_weeks"][0]

                    for category, _ in CATEGORIES:
                        st.session_state[f"score_{category}"] = int(item["scores"][category])

                    st.rerun()

            with b2:
                if st.button("REMOVE", use_container_width=True, disabled=not remove_customer):
                    st.session_state.pending_scores.pop(remove_customer, None)
                    st.rerun()

            with b3:
                if st.button("CLEAR BATCH", use_container_width=True):
                    st.session_state.pending_scores = {}
                    st.rerun()

            total_pending = len(st.session_state.pending_scores)
            avg_pending = (
                sum(sum(item["scores"].values()) for item in st.session_state.pending_scores.values())
                / (total_pending * MAX_TOTAL)
                * 100
            )

            m1, m2 = st.columns(2)
            m1.metric("Pending customers", total_pending)
            m2.metric("Average health", f"{avg_pending:.0f}%")

            if st.button(
                "ULTIMATE SAVE",
                type="primary",
                use_container_width=True,
            ):
                save_pending_batch()

        else:
            st.info("No pending scores yet. Select a customer, score it, then add it to the batch.")

        st.divider()

        st.download_button(
            "DOWNLOAD CUSTOMER_SCORING.XLSX",
            data=workbook_bytes(),
            file_name="Customer_Scoring.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# -------------------- CUSTOMER RANKING --------------------
with tab_rank:
    latest_year, latest_week = get_last_saved_period()

    top1, top2, top3 = st.columns([1.1, 1, 1])

    with top1:
        mode = st.selectbox(
            "View",
            options=["Latest Week", "Last 5 Weeks", "Custom"],
        )

    with top2:
        rank_region = st.selectbox(
            "Region",
            options=["All Regions"] + sorted(load_customer_master().keys()),
            key="rank_region",
        )

    with top3:
        rank_frequency = st.selectbox(
            "Frequency",
            options=["All", "Weekly", "Monthly"],
            key="rank_frequency",
        )

    custom_periods = None
    display_year = latest_year
    display_week = latest_week

    if mode == "Custom":
        c1, c2 = st.columns([1, 2])

        with c1:
            custom_year = st.number_input(
                "Custom Year",
                min_value=2020,
                max_value=2100,
                value=int(latest_year),
                step=1,
            )

        with c2:
            custom_weeks_text = st.text_input(
                "Weeks",
                placeholder="Examples: 30,31,34 or 28-34",
            )

        try:
            custom_periods = parse_custom_weeks(
                int(custom_year),
                custom_weeks_text,
            )
        except Exception:
            st.error("Use week input such as 30,31,34 or 28-34.")
            custom_periods = []

        if custom_periods:
            display_year, display_week = custom_periods[-1]

    rows, periods = prepare_visual_rows(
        latest_year,
        latest_week,
        mode,
        rank_region,
        rank_frequency,
        custom_periods=custom_periods,
    )

    if mode == "Custom" and not custom_periods:
        st.info("Enter one or more custom weeks to generate the ranking.")
    else:
        ranking_image = render_ranking(
            rows,
            display_year,
            display_week,
            mode,
            periods,
        )

        st.image(ranking_image, use_container_width=True)

        safe_mode = mode.replace(" ", "_")
        png_name = (
            f"Customer_Ranking_{safe_mode}_"
            f"{rank_region.replace(' ', '_')}_"
            f"{rank_frequency}.png"
        )

        st.download_button(
            "EXPORT PNG",
            data=image_to_png_bytes(ranking_image),
            file_name=png_name,
            mime="image/png",
            type="primary",
        )

        with st.expander("Category / Health explanation"):
            for category, tooltip in CATEGORIES:
                st.markdown(f"**{category}**")
                st.text(tooltip)

            st.markdown(
                """
                **Health bar**

                90–100% — Excellent  
                75–89% — Good  
                50–74% — Average  
                25–49% — Poor  
                0–24% — Critical
                """
            )

#endregion STREAMLIT APP
